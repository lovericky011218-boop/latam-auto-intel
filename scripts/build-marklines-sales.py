#!/usr/bin/env python3
"""Aggregate the supplied MarkLines workbook into the dashboard sales snapshot."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


TARGET_COUNTRIES = {
    "巴西", "阿根廷", "智利", "乌拉圭", "玻利维亚", "厄瓜多尔", "秘鲁",
    "意大利", "挪威", "荷兰", "西班牙", "法国", "以色列", "瑞典", "德国",
    "波兰", "比利时", "英国", "匈牙利", "澳大利亚", "新西兰", "泰国",
    "印度尼西亚", "马来西亚", "新加坡",
}


def aliases(*names: str, family: str) -> dict[str, str]:
    return {name: family for name in names}


MODEL_FAMILIES: dict[str, dict[str, str]] = {
    "比亚迪汽车(BYD)": {
        **aliases("Dolphin Mini/Dolphin Surf/ATTO 1 (Seagull)", "Seagull", family="BYD Seagull family"),
        **aliases("Dolphin", family="Dolphin"),
        **aliases("Yuan Pro", family="Yuan Pro"),
        **aliases("ATTO 2 (Yuan UP)", "Yuan UP", family="BYD Atto 2 / Yuan Up"),
        **aliases("Yuan PLUS / ATTO 3", family="BYD Atto 3 / Yuan Plus"),
        **aliases("Song Pro", "Sealion 05", family="BYD Song Pro / Sealion 5"),
        **aliases("Song PLUS", "Seal U", "Seal U/Sealion 6 (Song Plus)", "Sealion 6", family="BYD Song Plus / Seal U / Sealion 6"),
        **aliases("BYD King", "Destroyer 05/BYD King/BYD Chazor (Haibao/Seal 05)", family="King"),
        **aliases("Haibao/Seal", family="Seal"),
        **aliases("Haibao/Seal 06", family="BYD Seal 6 family"),
        **aliases("Sealion 07", family="Sealion 7"),
        **aliases("Tang", family="Tang"),
        **aliases("ATTO 8/Sealion 8 (Tang L)", "Sealion 08", family="BYD Atto 8 / Sealion 8"),
        **aliases("Shark / Shark 6", family="Shark"),
        **aliases("BYD M6", "eMax7", family="M6"),
        **aliases("Denza B5", family="Denza B5"),
        **aliases("Denza D9", family="Denza D9"),
    },
    "奇瑞集团 (Chery)": {
        **aliases("Tiggo 2", family="Tiggo 2 Pro"),
        **aliases("Tiggo 4", "Tiggo 5X", family="Chery Tiggo 4 / Tiggo 5X"),
        **aliases("Tiggo 7", family="Tiggo 7"),
        **aliases("Tiggo 8", "C8 (Tiggo 8)", family="Tiggo 8 Pro"),
        **aliases("Tiggo 9", family="Tiggo 9"),
        **aliases("Arrizo 5", family="Arrizo 5 Pro"),
        **aliases("Arrizo 8", family="Arrizo 8 CSH"),
        **aliases("Chery M7", family="M7"),
        **aliases("HIMLA", family="Himla"),
        **aliases("OMODA 5", "OMODA 5/C5", family="Omoda 5 family"),
        **aliases("OMODA 7", "OMODA 7/C7", "Omoda C7", family="Omoda 7"),
        **aliases("OMODA 9", family="Omoda 9"),
        **aliases("Jaecoo 5", "Jaecoo 5/J5", family="Jaecoo 5"),
        **aliases("Jaecoo 6", family="Jaecoo 6"),
        **aliases("Jaecoo 7", "Jaecoo 7/J7", family="Jaecoo 7"),
        **aliases("Jaecoo 8", "Jaecoo 8/J8", family="Jaecoo 8"),
        **aliases("EXEED LX", family="Exeed LX"),
        **aliases("EXEED TXL", family="Exeed TXL"),
        **aliases("JETOUR X50", family="Jetour X50"),
        **aliases("JETOUR Dasheng", "JETOUR Dashing (Dasheng)", family="Jetour Dashing"),
        **aliases("JETOUR X70", family="Jetour X70"),
        **aliases("JETOUR T1 (JETOUR Ziyouzhe)", "JETOUR T-1", "JETOUR Shanhai T1", family="Jetour T1"),
        **aliases("JETOUR T2 (JETOUR T-1)", "JETOUR T-2", "JETOUR Shanhai T2", family="Jetour T2"),
    },
    "吉利控股集团(Geely)": {
        **aliases("Geely GX3 PRO", "GX3 Pro", family="GX3 Pro"),
        **aliases("Emgrand", "Emgrand / Emgrand EC7", family="Emgrand"),
        **aliases("Coolray", family="Coolray"),
        **aliases("Cityray", family="Cityray"),
        **aliases("Geely Starray", "StarRay", family="Starray"),
        **aliases("Okavango", family="Okavango"),
        **aliases("EX2/Star Wish (Geely Xingyuan)", family="EX2"),
        **aliases("Geely E5/EX5/e.MAS 7 (Galaxy E5)", "Geely EX5", "EX5 EM-i", family="Geely EX5 family"),
        **aliases("ZEEKR X", family="Zeekr X"),
        **aliases("ZEEKR 001", family="Zeekr 001"),
        **aliases("ZEEKR 009", family="Zeekr 009"),
        **aliases("ZEEKR 7X", family="Zeekr 7X"),
        **aliases("ZEEKR 7GT (ZEEKR 007)", family="Zeekr 7GT"),
        **aliases("LYNK & CO 01", family="Lynk & Co 01"),
        **aliases("LYNK & CO 02", family="Lynk & Co 02"),
        **aliases("LYNK & CO 03", family="Lynk & Co 03+"),
        **aliases("LYNK & CO 06", family="Lynk & Co 06"),
        **aliases("LYNK & CO 08 EM-P", family="Lynk & Co 08"),
        **aliases("LYNK & CO 09", family="Lynk & Co 09"),
    },
    "长城汽车股份有限公司 (GWM)": {
        **aliases("ORA/ORA 03/Good Cat/Funky Cat (Haomao)", family="Ora 03"),
        **aliases("Jolion (F&L)", "Jolion Pro", family="Haval Jolion"),
        **aliases("H6", "Haval H6", family="Haval H6"),
        **aliases("Haval H7", family="Haval H7"),
        **aliases("Haval H9", family="Haval H9"),
        **aliases("Tank 300", family="Tank 300"),
        **aliases("Tank 500", family="Tank 500"),
        **aliases("Cannon", "GW Poer", "GWM Ute", "P-Series", "Poer", family="Poer"),
        **aliases("Wingle 5", family="Wingle 5"),
        **aliases("GW Wingle 7", family="Wingle 7"),
        **aliases("WEY 03 / Coffee 02", "WEY 03", "Coffee 02", family="WEY 03"),
        **aliases("WEY 05 / Coffee 01", "WEY 05", "Coffee 01", family="WEY 05"),
    },
    "东风汽车公司": {
        **aliases("Nammi Box/Dongfeng Box (NAMMI 01)", "NAMMI 01", family="Dongfeng BOX / Nammi"),
        **aliases("Vigo (NAMMI 06)", "NAMMI 06", family="Vigo"),
        **aliases("E70", "Aeolus E70", family="E70"),
        **aliases("Mage (Haohan/Aeolus L7)", family="Dongfeng Mage family"),
        **aliases("Huge", "Huge (Haoji)", family="Huge"),
        **aliases("Paladin", family="Paladin"),
        **aliases("Rich", family="Dongfeng Rich family"),
        **aliases("Dongfeng Z9", family="Z9"),
        **aliases("FREE", family="VOYAH Free"),
        **aliases("Courage (Zhiyin)", family="VOYAH Courage"),
        **aliases("Dream", family="VOYAH Dream"),
        **aliases("Voyah Passion (Zhuiguang)", family="VOYAH Passion"),
    },
    "零跑汽车 (Leapmotor)": {
        **aliases("T03", family="T03"),
        **aliases("B10", family="B10"),
        **aliases("C10", family="C10"),
    },
    "长安汽车集团(Changan/Chana)": {
        **aliases("UNI-T", family="UNI-T"),
        **aliases("CS55 PLUS", family="CS55 Plus"),
        **aliases("Eado", family="Eado Plus"),
        **aliases("Lumin", family="Lumin"),
        **aliases("CS75 PLUS", family="CS75 Plus"),
        **aliases("Alsvin", family="Alsvin"),
        **aliases("CS35 PLUS", family="CS35 Plus"),
        **aliases("CS15", family="CS15"),
        **aliases("UNI-K", family="UNI-K"),
        **aliases("欧尚 (Oushang) X7", family="X7 Plus"),
        **aliases("Hunter (Changan/Chana)", family="Hunter"),
        **aliases("F70", family="F70"),
        **aliases("S05", family="Deepal S05"),
        **aliases("S07", family="Deepal S07"),
        **aliases("G318", family="Deepal G318"),
        **aliases("Avatr 11", family="AVATR 11"),
        **aliases("Avatr 07", family="AVATR 07"),
    },
    "小鹏汽车 (XPeng)": {
        **aliases("XPeng G6", family="XPENG G6"),
        **aliases("XPeng G9", family="XPENG G9"),
        **aliases("XPeng P7+", family="XPENG P7+"),
        **aliases("XPeng X9", family="XPENG X9"),
    },
    "蔚来汽车 (NIO)": {
        **aliases("NIO ET5", family="NIO ET5 family"),
        **aliases("NIO EL6", family="NIO EL6"),
        **aliases("NIO EL8", "NIO ES8", family="NIO EL8"),
        **aliases("Firefly", family="firefly"),
    },
    "广汽集团": {
        **aliases("AionUT", family="AION UT"),
        **aliases("AionV", family="AION V"),
        **aliases("AionY", family="AION Y / Y Plus"),
        **aliases("Aion ES", family="AION ES"),
        **aliases("Hyptec HT", family="HYPTEC HT"),
        **aliases("EMZOOM / Trumpchi GS3", "Trumpchi GS3", family="GAC GS3 family"),
        **aliases("Trumpchi GS4", family="GAC GS4 family"),
        **aliases("Emkoo", family="EMKOO"),
        **aliases("Trumpchi GS8", family="GS8"),
        **aliases("Trumpchi Empow", family="EMPOW"),
        **aliases("Trumpchi M8", "Trumpchi GM8 Qiankun", family="GAC M8 / GN8"),
        **aliases("GN6", family="M6 Pro"),
        **aliases("Trumpchi S7", family="S7"),
    },
    "其他中小整车集团": {
        **aliases("S400", family="Chery Tiggo 4 / Tiggo 5X"),
        **aliases("S700", family="Tiggo 7"),
        **aliases("S800", family="Tiggo 8 Pro"),
        **aliases("S900", family="Tiggo 9"),
    },
}

ENERGY = {"BEV": "纯电", "PHEV": "插混", "HEV": "混动", "ICE": "燃油", "FCV": "氢能"}


def number(value: object) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def build(workbook: Path, output: Path) -> None:
    sheet = load_workbook(workbook, read_only=True, data_only=True)["Sheet1"]
    totals: dict[tuple[str, str, str], list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
    rows_seen: set[tuple[str, str, str]] = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        country, group, model, energy = row[0], row[2], row[7], row[8]
        if country not in TARGET_COUNTRIES or group not in MODEL_FAMILIES or model not in MODEL_FAMILIES[group]:
            continue
        family = MODEL_FAMILIES[group][model]
        energy_name = ENERGY.get(str(energy))
        if not energy_name:
            continue
        key = (str(country), family, energy_name)
        totals[key][0] += sum(number(value) for value in row[10:22])
        totals[key][1] += sum(number(value) for value in row[22:34])
        totals[key][2] += sum(number(value) for value in row[34:38])
        rows_seen.add(key)

    payload = [
        {
            "country": country,
            "model": model,
            "energy": energy,
            "y2024": round(values[0]),
            "y2025": round(values[1]),
            "y2026": round(values[2]),
        }
        for (country, model, energy), values in sorted(totals.items())
        if (country, model, energy) in rows_seen
    ]
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    output.write_text(
        "// Generated by scripts/build-marklines-sales.py from the user-supplied MarkLines workbook.\n"
        "export type MarklinesSalesRow = { country:string; model:string; energy:string; y2024:number; y2025:number; y2026:number };\n"
        f"export const marklinesSales: MarklinesSalesRow[] = {serialized};\n"
        "export const marklinesSalesPeriod = \"2024-01 至 2026-04\";\n"
        "export const marklinesSalesSource = \"MarkLines｜2024-2026最新(1).xlsx｜Sheet1\";\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(payload)} aggregates to {output}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    build(args.workbook, args.output)
